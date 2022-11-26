import csv
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit

matplotlib.rcParams.update({'font.size': 8})
maxLen = [0]
years = []
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

DynamicsOfSalaryLevelsByYears = {}
DynamicsOfTheNumbeOfVacanciesByYears = {}
DynamicsOfTheLevelOfSalariesByYearsForChosenProfession = {}
DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession = {}
SalaryLevelsByCity = {}
ShareVacanciesByCity = {}
sortedShareVacanciesByCity = {}
sortedSalaryLevelsByCity = {}
lists_by_years = []
lists_by_city1 = []
lists_by_city2 = []


def yearCheck(years):
    for year in years:
        if year not in DynamicsOfTheLevelOfSalariesByYearsForChosenProfession.keys():
            DynamicsOfTheLevelOfSalariesByYearsForChosenProfession[year] = 0

        if year not in DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession.keys():
            DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession[year] = 0


class Report:
    def generate_excel(self, sortedSalaryLevelsByCity, sortedShareVacanciesByCity):
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика по годам"
        ws1 = wb.create_sheet("Mysheet", 1)
        ws1.title = "Статистика по городам"

        ws['A1'] = "Год"
        ws['A1'].font = Font(bold=True)
        ws['B1'] = "Средняя зарплата"
        ws['B1'].font = Font(bold=True)
        ws['C1'] = f"Средняя зарплата - {inputValue.name_profession}"
        ws['C1'].font = Font(bold=True)
        ws['D1'] = "Количество вакансий"
        ws['D1'].font = Font(bold=True)
        ws['E1'] = f"Количество вакансий - {inputValue.name_profession}"
        ws['E1'].font = Font(bold=True)

        dict_len_colum1 = {
            'A': len("Год"),
            'B': len("Средняя зарплата"),
            "C": len(f"Средняя зарплата - {inputValue.name_profession}"),
            "D": len("Количество вакансий"),
            "E": len(f"Количество вакансий - {inputValue.name_profession}")
        }

        ws1['A1'] = "Город"
        ws1['A1'].font = Font(bold=True)
        ws1['B1'] = "Уровень зарплат"
        ws1['B1'].font = Font(bold=True)
        ws1.column_dimensions["C"].width = 3
        ws1['D1'] = "Город"
        ws1['D1'].font = Font(bold=True)
        ws1['E1'] = "Доля вакансий"
        ws1['E1'].font = Font(bold=True)

        dict_len_colum2 = {
            'A': len("Город"),
            'B': len("Уровень зарплат"),
            "D": len("Город"),
            "E": len("Доля вакансий")
        }

        for sumbol in ['A', 'B', "C", "D", "E"]:
            ws[sumbol + str(1)].border = thin_border
            if sumbol == 'A':
                fillColomYear(ws, years, sumbol, dict_len_colum1)
                ws.column_dimensions[sumbol].width = maxLen[0] + 1
                maxLen[0] = 0
            else:
                fillColomYearValue(ws, years, sumbol, dict_func[sumbol], dict_len_colum1)
                ws.column_dimensions[sumbol].width = maxLen[0] + 1
                maxLen[0] = 0

        for sumbol in ['A', 'B', "D", "E"]:
            ws1[sumbol + str(1)].border = thin_border
            if sumbol == 'A' or sumbol == 'D':
                if sumbol == 'A':
                    fillColomCity(ws1, sortedSalaryLevelsByCity, sumbol, dict_len_colum2)
                else:
                    fillColomCity(ws1, sortedShareVacanciesByCity, sumbol, dict_len_colum2)
                ws1.column_dimensions[sumbol].width = maxLen[0] + 3
                maxLen[0] = 0
            else:
                if sumbol == 'B':
                    fillColomCityValue(ws1, sortedSalaryLevelsByCity, sumbol, dict_len_colum2)
                else:
                    fillColomCityValue(ws1, sortedShareVacanciesByCity, sumbol, dict_len_colum2, ' %')
                ws1.column_dimensions[sumbol].width = maxLen[0] + 3
                maxLen[0] = 0

        for i in range(2, len(years) + 2):
            list = []
            for sumbol in ['A', 'B', "C", "D", "E"]:
                list.append(int(ws[sumbol + str(i)].value))
            lists_by_years.append(list)

        for i in range(2,12):
            list = []
            for sumbol in ['A', 'B']:
                list.append(ws1[sumbol + str(i)].value)
            lists_by_city1.append(list)

        for i in range(2,12):
            list =  []
            for sumbol in ["D", "E"]:
                list.append(ws1[sumbol + str(i)].value)
            lists_by_city2.append(list)
        if inputValue.jobs_or_statistics == 'Вакансии':
            wb.save('report.xlsx')

    def createPDF(self):

        env = Environment(loader=FileSystemLoader(""))
        template = env.get_template("sample.html")

        pdf_template = template.render({'name': inputValue.name_profession, 'lists_by_years': lists_by_years, "lists_by_city1": lists_by_city1,
                                        "lists_by_city2": lists_by_city2, "image_file": 'saved_figure.png'})

        config = pdfkit.configuration(wkhtmltopdf=r'D:\Old Data\Program\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'Report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def generate_image(self):
        creatFirstGraf()


def creatFirstGraf():
    years = []
    city = []
    cityShareVacancies = []
    list = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    list6 = []
    list7 = []

    for key, value in DynamicsOfSalaryLevelsByYears.items():
        years.append(int(key))
        list.append(int(value))
    for key, value in DynamicsOfTheLevelOfSalariesByYearsForChosenProfession.items():
        list2.append(int(value))
    for key, value in DynamicsOfTheNumbeOfVacanciesByYears.items():
        list3.append(int(value))
    for key, value in DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession.items():
        list4.append(int(value))
    for key, value in sortedSalaryLevelsByCity.items():
        city.append(key)
        list5.append(int(value))
    sum = 0
    cityShareVacancies.append("Другие")
    for key, value in sortedShareVacanciesByCity.items():
        list6.append(value)
        cityShareVacancies.append(key)
    for key, value in ShareVacanciesByCity.items():
        if key not in sortedShareVacanciesByCity:
            sum += value * 100
    list7.append(sum)
    list7 = list7 + list6

    city.reverse()
    list5.reverse()

    x = np.arange(len(years))
    y = np.arange(len(sortedShareVacanciesByCity))

    width = 0.35

    fig, axs = plt.subplots(2, 2)
    axs[0, 0].bar(x - width / 2, list, width, label='Средняя з/п')
    axs[0, 0].bar(x + width / 2, list2, width, label=f'з/п {inputValue.name_profession}')

    axs[0, 1].bar(x - width / 2, list3, width, label='Количество вакансий')
    axs[0, 1].bar(x + width / 2, list4, width, label=f'Количество вакансий \n{inputValue.name_profession}')

    axs[1, 0].barh(y, list5, width)

    axs[1, 1].pie(list7, labels=cityShareVacancies, textprops={'fontsize': 6})

    axs[0, 0].set_title('Уровень зарплат по годам')
    axs[0, 0].set_xticks(x, years, rotation=90)
    axs[0, 0].legend()

    axs[0, 1].set_title('Количество вакансий по годам')
    axs[0, 1].set_xticks(x, years, rotation=90)
    axs[0, 1].legend()

    axs[1, 0].set_title('Уровень зарплат по городам')
    axs[1, 0].set_yticks(y, city, size=6)

    axs[1, 1].set_title('Доля вакансий по городам')
    fig.tight_layout()

    plt.savefig('D:\\PythonProject\\report\\saved_figure.png')
    if inputValue.jobs_or_statistics == 'Статистика':
        plt.show()


def fillColomYear(ws, years, sumbol, dict_len_colum):
    maxLen[0] = dict_len_colum[sumbol]
    for num in range(2, len(years) + 2):
        ws[sumbol + str(num)].border = thin_border
        ws[sumbol + str(num)] = years[num - 2]
        if len(str(years[num - 2])) > maxLen[0]:
            maxLen[0] = len(str(years[num - 2]))


def fillColomYearValue(ws, years, sumbol, dict, dict_len_colum):
    maxLen[0] = dict_len_colum[sumbol]
    for num in range(2, len(years) + 2):
        ws[sumbol + str(num)].border = thin_border
        ws[sumbol + str(num)] = dict[years[num - 2]]
        if len(str(dict[years[num - 2]])) > maxLen[0]:
            maxLen[0] = len(str(years[num - 2]))


def fillColomCity(ws1, dict, sumbol, dict_len_colum):
    maxLen[0] = dict_len_colum[sumbol]
    i = 2
    for name in dict:
        if i == 12:
            break
        ws1[sumbol + str(i)].border = thin_border
        ws1[sumbol + str(i)] = name
        if len(name) > maxLen[0]:
            maxLen[0] = len(name)
        i += 1


def fillColomCityValue(ws1, dict, sumbol, dict_len_colum, special_character=""):
    maxLen[0] = dict_len_colum[sumbol]
    i = 2
    for name in dict:
        if i == 12:
            break
        ws1[sumbol + str(i)].border = thin_border
        if special_character != '':
            dict[name] = round(dict[name] * 100, 2)
        ws1[sumbol + str(i)] = str(dict[name]) + special_character
        if len(str(dict[name]) + special_character) > maxLen[0]:
            maxLen[0] = len(str(dict[name]) + special_character)
        i += 1


dict_func = {
    'B': DynamicsOfSalaryLevelsByYears,
    "C": DynamicsOfTheLevelOfSalariesByYearsForChosenProfession,
    "D": DynamicsOfTheNumbeOfVacanciesByYears,
    "E": DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession
}


class InputConect:
    file_name = ""
    name_profession = ""
    jobs_or_statistics = ''

    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.name_profession = input("Введите название профессии: ")
        self.jobs_or_statistics = input("Вакансии или Статистика(нужно вписать Вакансии или Статистика): ")


class DataSet:
    file_name = ""

    def csv_reader(self, file_name):
        self.file_name = file_name
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as csvfile:
            data = csv.reader(csvfile, delimiter=',')
            list_row = []
            for row in data:
                list_row.append(row)
            if len(list_row) == 0:
                print(("Пустой файл"))
                sys.exit()
            headers = list_row.pop(0)
            return headers, list_row

    def csv_filter(self, headers, list_row):
        if len(list_row) == 0:
            print("Нет данных")
            sys.exit()
        countVacancies = 0
        for row in list_row:
            dict = {}
            if len(headers) == len(row) and row.count('') == 0:
                for i in range(len(row)):
                    dict[headers[i]] = row[i]

                completingDictionary(DynamicsOfSalaryLevelsByYears, dict, dict['published_at'][:4])
                completingDictionary(SalaryLevelsByCity, dict, dict['area_name'])
                countCompletingDictionary(DynamicsOfTheNumbeOfVacanciesByYears, dict, dict['published_at'][:4])
                countCompletingDictionary(ShareVacanciesByCity, dict, dict['area_name'])

                if inputValue.name_profession in dict['name']:
                    completingDictionary(DynamicsOfTheLevelOfSalariesByYearsForChosenProfession, dict,
                                         dict['published_at'][:4])
                    countCompletingDictionary(DynamicsOfTheNumbeOfVacanciesByYearsForChosenProfession, dict,
                                              dict['published_at'][:4])
                countVacancies += 1

                if dict['published_at'][:4] not in years:
                    years.append(dict['published_at'][:4])
        convertListAndOutput(countVacancies)


def completingDictionary(dictCompleting, dict, name):
    if name not in dictCompleting.keys():
        dictCompleting[name] = [
            int((float(dict['salary_from']) + float(dict['salary_to'])) / 2) * currency_to_rub[dict['salary_currency']],
            1]
    else:
        dictCompleting[name] = [
            int(dictCompleting[name][0]) +
            int((float(dict['salary_from']) + float(dict['salary_to'])) * currency_to_rub[dict['salary_currency']] / 2),
            int(dictCompleting[name][1]) + 1]


def countCompletingDictionary(dictCompleting, dict, name):
    if name not in dictCompleting.keys():
        dictCompleting[name] = 1
    else:
        dictCompleting[name] = int(dictCompleting[name]) + 1


def getValueFromArray(list, countVacancies=0, rounded=0):
    for i in list:
        if rounded == 0:
            list[i] = int(list[i][0] / list[i][1])
        else:
            list[i] = round(int(list[i]) / countVacancies, rounded)


def convertListAndOutput(countVacancies):
    getValueFromArray(DynamicsOfSalaryLevelsByYears)
    getValueFromArray(DynamicsOfTheLevelOfSalariesByYearsForChosenProfession)
    getValueFromArray(ShareVacanciesByCity, countVacancies, 4)

    sortedVacanciesKeys = sorted(ShareVacanciesByCity, key=ShareVacanciesByCity.get, reverse=True)
    count = 0
    for city in sortedVacanciesKeys:
        if (count >= 10 or ShareVacanciesByCity[city] < 0.01):
            break
        count += 1
        sortedShareVacanciesByCity[city] = ShareVacanciesByCity[city]

    count = 0
    for i in SalaryLevelsByCity:
        SalaryLevelsByCity[i] = int(SalaryLevelsByCity[i][0] / SalaryLevelsByCity[i][1])

    sortedSalaryKeys = sorted(SalaryLevelsByCity, key=SalaryLevelsByCity.get, reverse=True)

    for city in sortedSalaryKeys:
        if ShareVacanciesByCity[city] >= 0.01:
            if (count >= 10):
                break
            count += 1
            sortedSalaryLevelsByCity[city] = SalaryLevelsByCity[city]

    yearCheck(years)
    repotr = Report()

    repotr.generate_excel(sortedSalaryLevelsByCity, sortedShareVacanciesByCity)
    repotr.generate_image()
    repotr.createPDF()


inputValue = InputConect()
if inputValue.jobs_or_statistics not in ['Вакансии', 'Статистика']:
    print("Ввод некорректен")
    sys.exit()

data_set = DataSet()
headers, list_row = data_set.csv_reader(inputValue.file_name)
data_set.csv_filter(headers, list_row)